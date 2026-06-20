from __future__ import annotations

import shutil
import tempfile
import unittest
from datetime import date
from pathlib import Path
from zipfile import ZipFile

from openpyxl import load_workbook

from utiles.bnz_csv import AccountResolver, StatementCoverageError, csv_paths, validate_statement_coverage
from utiles.ai_classifier import AITransactionContext, SuggestedRule, append_rules, dedupe_contexts, likely_sensitive_payee, normalize_ai_classification
from utiles.converter import BnzCsvToIcostConverter
from utiles.models import Classification
from utiles.rules import RuleEngine


class ConverterIntegrationTest(unittest.TestCase):
    def test_converter_outputs_icost_workbook_and_unknown_file(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            root = Path(tmpdir)
            input_dir = root / "bnz_statements"
            config_dir = root / "config"
            output_dir = root / "output"
            shutil.copytree(Path("tests/data/bnz_statements"), input_dir)
            shutil.copytree(Path("tests/data/config"), config_dir)

            output_file, unknown_file, row_count, unknown_count = BnzCsvToIcostConverter(
                input_dir=input_dir,
                output_dir=output_dir,
                config_dir=config_dir,
                date_from=date(2026, 1, 1),
                date_to=date(2026, 1, 31),
            ).run()

            self.assertEqual(row_count, 4)
            self.assertEqual(unknown_count, 0)
            self.assertTrue(output_file.exists())
            self.assertTrue(unknown_file.exists())

            workbook = load_workbook(output_file, data_only=True, read_only=True)
            rows = list(workbook.active.iter_rows(values_only=True))
            self.assertEqual(rows[0], ("日期", "类型", "金额", "一级分类", "二级分类", "账户1", "账户2", "备注", "货币", "标签"))
            self.assertIn(("2026年01月02日 12:00:00", "转账", 100, "其他", None, "Main Account", "Savings"), [row[:7] for row in rows[1:]])
            self.assertIn(("2026年01月04日 12:00:00", "转账", 1000, "其他", None, "Savings", "Term Deposit 3333333-01"), [row[:7] for row in rows[1:]])

            with ZipFile(output_file) as archive:
                sheet_xml = archive.read("xl/worksheets/sheet1.xml").decode("utf-8")
                self.assertIsNone(archive.testzip())
                self.assertIn("xl/sharedStrings.xml", archive.namelist())
                self.assertNotIn("inlineStr", sheet_xml)

            unknown_workbook = load_workbook(unknown_file, data_only=True, read_only=True)
            self.assertEqual(unknown_workbook.active.max_row, 1)

    def test_statement_filename_coverage_is_required(self) -> None:
        paths = csv_paths(Path("tests/data/bnz_statements"))
        resolver = AccountResolver.from_csv(Path("tests/data/config/accounts.csv"))

        validate_statement_coverage(paths, resolver, date(2026, 1, 1), date(2026, 1, 31))
        with self.assertRaises(StatementCoverageError):
            validate_statement_coverage(paths, resolver, date(2026, 1, 1), date(2026, 2, 1))

    def test_default_rules_are_loaded_from_csv(self) -> None:
        rules = RuleEngine.load(
            local_rules_file=Path("tests/data/config/missing_local_rules.csv"),
            default_rules_file=Path("tests/data/config/default_rules.csv"),
        )

        self.assertTrue(any(rule.match_text == "PAK N SAVE" for rule in rules.rules))

    def test_ai_rule_storage_and_privacy_filter(self) -> None:
        self.assertTrue(likely_sensitive_payee("PERSON,NAME"))
        self.assertTrue(likely_sensitive_payee("John Smith"))
        self.assertTrue(likely_sensitive_payee("Person One&Person Two"))
        self.assertFalse(likely_sensitive_payee("PAK N SAVE"))
        self.assertFalse(likely_sensitive_payee("WASHWORLD WAIRAU"))
        self.assertFalse(likely_sensitive_payee("SUBWAY SUNNYBRAE"))
        self.assertFalse(likely_sensitive_payee("GetYourGuide Tickets"))
        self.assertFalse(likely_sensitive_payee("CHARLIE'S TEA"))

        self.assertEqual(normalize_ai_classification("支出", "生活费", "饮品"), Classification("支出", "生活费", "外食"))
        self.assertEqual(normalize_ai_classification("支出", "旅游", "门票/活动"), Classification("支出", "旅游", ""))
        self.assertEqual(normalize_ai_classification("收入", "旅游", ""), Classification("收入", "旅游", ""))

        context = AITransactionContext(
            payee="John Smith",
            particulars="Lunch",
            code="Cafe",
            reference="",
            tran_type="BP",
            direction="收入",
        )
        self.assertEqual(context.to_payload()["payee"], "[private_payee]")
        self.assertEqual(context.to_payload()["particulars"], "Lunch")
        self.assertEqual(len(dedupe_contexts([context, context])), 1)
        self.assertEqual(
            dedupe_contexts([AITransactionContext("John Smith", "", "", "", "BP", "收入")]),
            [],
        )

        with tempfile.TemporaryDirectory() as tmpdir:
            rules_file = Path(tmpdir) / "local_rules.csv"
            added = append_rules(
                rules_file,
                [
                    SuggestedRule(
                        match_text="KFC",
                        classification=Classification("支出", "生活费", "外食"),
                        reason="known fast food merchant",
                        confidence=0.95,
                    )
                ],
            )

            self.assertEqual(added, 1)
            self.assertIn("KFC", rules_file.read_text(encoding="utf-8-sig"))


if __name__ == "__main__":
    unittest.main()
