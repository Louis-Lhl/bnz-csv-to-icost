from __future__ import annotations

import csv
import logging
from dataclasses import dataclass
from decimal import Decimal
from pathlib import Path

from openpyxl import load_workbook

from .models import Classification, Transaction, transaction_key, tx_key

LOGGER = logging.getLogger(__name__)

VALID_TYPES = {"支出", "收入", "转账"}
EXPENSE_CATEGORIES = {
    "生活费",
    "购物",
    "服饰",
    "日用",
    "数码",
    "美妆",
    "护肤",
    "应用软件",
    "住房",
    "交通",
    "娱乐",
    "医疗",
    "通讯",
    "汽车",
    "学习",
    "办公",
    "运动",
    "社交",
    "人情",
    "育儿",
    "宠物",
    "旅行",
    "旅游",
    "度假",
    "烟酒",
    "其他",
}
INCOME_CATEGORIES = {"工资", "奖金", "加班", "福利", "公积金", "红包", "兼职", "副业", "退税", "投资", "意外收入", "生活费", "应用软件", "其他"}
SHORT_CATEGORY_MAP = {
    "外食": ("生活费", "外食"),
    "三餐": ("生活费", "三餐"),
    "零食": ("生活费", "零食"),
    "水果": ("生活费", "水果"),
    "蔬菜": ("生活费", "蔬菜"),
    "房租": ("住房", "房租"),
}


@dataclass(frozen=True)
class Rule:
    match_text: str
    classification: Classification
    priority: int = 100
    direction: str = ""
    bank_type: str = ""
    amount: Decimal | None = None

    def matches(self, tx: Transaction) -> bool:
        text = tx.particulars.upper()
        if self.direction and self.direction != tx.direction:
            return False
        if self.bank_type and self.bank_type != tx.bank_type:
            return False
        if self.amount is not None and self.amount != tx.amount:
            return False
        return self.match_text.upper() in text


DEFAULT_RULES = [
    Rule("AUCKLAND TRANSPORT", Classification("支出", "交通")),
    Rule("AT PUBLIC TRANSPORT", Classification("支出", "交通")),
    Rule("SNAPPER SERVICES", Classification("支出", "交通")),
    Rule("DIDI_NZ", Classification("支出", "交通")),
    Rule("UBER", Classification("支出", "交通")),
    Rule("AIRPORT PARKING", Classification("支出", "交通")),
    Rule("PAK N SAVE", Classification("支出", "生活费")),
    Rule("WOOLWORTHS", Classification("支出", "生活费")),
    Rule("NEW WORLD", Classification("支出", "生活费")),
    Rule("FOUR SQUARE", Classification("支出", "生活费")),
    Rule("DH SUPERMARKET", Classification("支出", "生活费")),
    Rule("GOLDEN APPLE", Classification("支出", "生活费")),
    Rule("YI CART", Classification("支出", "生活费")),
    Rule("PRIME FOODS", Classification("支出", "生活费")),
    Rule("POKENO BACON", Classification("支出", "生活费")),
    Rule("DA HUA SUPERMARK", Classification("支出", "生活费")),
    Rule("WANG MART", Classification("支出", "生活费")),
    Rule("CHOWON BUTCHERY", Classification("支出", "生活费")),
    Rule("LINK VEGE", Classification("支出", "生活费")),
    Rule("FARRO FRESH", Classification("支出", "生活费")),
    Rule("SIMPLY FRESH", Classification("支出", "生活费")),
    Rule("APPLE.COM/BILL", Classification("支出", "应用软件")),
    Rule("GITHUB", Classification("支出", "应用软件")),
    Rule("OPENAI", Classification("支出", "应用软件")),
    Rule("GOOGLE SURFSHARK", Classification("支出", "应用软件")),
    Rule("THEVERGE.COM", Classification("支出", "应用软件")),
    Rule("ONE NZ", Classification("支出", "通讯")),
    Rule("POWERSHOP", Classification("支出", "住房", "水电费")),
    Rule("MOBIL", Classification("支出", "汽车")),
    Rule("BP CONNECT", Classification("支出", "汽车")),
    Rule("Z ENERGY", Classification("支出", "汽车")),
    Rule("Z SUNNYBRAE", Classification("支出", "汽车")),
    Rule("NZTA", Classification("支出", "汽车")),
    Rule("NZ TRANSPORT AGENCY", Classification("支出", "汽车")),
    Rule("AA INSURANCE", Classification("支出", "汽车")),
    Rule("L & Y MOTORS", Classification("支出", "汽车")),
    Rule("CHEMIST WAREHOUSE", Classification("支出", "医疗")),
    Rule("FULL HEALTH", Classification("支出", "医疗")),
    Rule("CITYFITNESS", Classification("支出", "运动")),
    Rule("AHSC BADMINTON", Classification("支出", "运动")),
    Rule("REBEL", Classification("支出", "运动")),
    Rule("8 BILLIARDS", Classification("支出", "运动")),
    Rule("ANIMATES", Classification("支出", "宠物")),
    Rule("BRISCOES", Classification("支出", "日用")),
    Rule("BUNNINGS", Classification("支出", "日用")),
    Rule("MITRE 10", Classification("支出", "日用")),
    Rule("THE WAREHOUSE", Classification("支出", "日用")),
    Rule("IKEA", Classification("支出", "日用")),
    Rule("KINGS PLANT BARN", Classification("支出", "日用")),
    Rule("3 DOLLAR JAPAN", Classification("支出", "日用")),
    Rule("NOEL LEEMING", Classification("支出", "数码")),
    Rule("PB TECHNOLOGIES", Classification("支出", "数码")),
    Rule("HARVEYNORMA", Classification("支出", "数码")),
    Rule("HIKOCO", Classification("支出", "美妆")),
    Rule("MECCA", Classification("支出", "美妆")),
    Rule("DAIKOKU COSMETICS", Classification("支出", "美妆")),
    Rule("HAIR SALON", Classification("支出", "护肤")),
    Rule("HAIR CUT", Classification("支出", "护肤")),
    Rule("THE BOTTLE-O", Classification("支出", "烟酒")),
    Rule("BLACK BULL LIQUOR", Classification("支出", "烟酒")),
    Rule("GARAGE PROJECT", Classification("支出", "烟酒")),
    Rule("AFTERPAY", Classification("支出", "购物")),
    Rule("DEPT OF INTERNAL AFF", Classification("支出", "其他")),
    Rule("EVENT CINEMAS", Classification("支出", "娱乐")),
    Rule("UGG EXPRESS", Classification("支出", "服饰")),
    Rule("AIR NEW ZEALAND", Classification("支出", "旅游")),
    Rule("AELIA DUTY FREE", Classification("支出", "旅游")),
    Rule("CATHEDRAL COVE", Classification("支出", "旅游")),
    Rule("I.R.D.", Classification("收入", "退税"), direction="收入"),
    Rule("GROSS CR INTEREST", Classification("收入", "投资"), direction="收入"),
    Rule("COCA-COLA EP NZ", Classification("支出", "生活费", "零食")),
]

FOOD_RULES = [
    "AZABU",
    "BAKERY",
    "BRILLVIC LIMITED",
    "BURGER",
    "CAFE",
    "COFFEE",
    "CINNABON",
    "CHUBBY BOY",
    "DAILY BREAD",
    "DWEJI",
    "GREAT TONG GI",
    "H&H RETAILING",
    "HELLO MISTER",
    "ITALIAN IN THE VILLA",
    "KFC",
    "KITCHEN",
    "KHAO GAENG",
    "LA TEXICAN",
    "LUKES KITCHEN",
    "MAD POT",
    "MEDI DELI",
    "MCDONALD",
    "MIGHTY HOTDOG",
    "MOJO THE SUMMIT",
    "NAVAT",
    "PHO",
    "RESTAURANT",
    "RINGAWERA",
    "SAIGON PHO",
    "SUSHI",
    "TAIKIN",
    "TAKAPUNA BEACHSIDE",
    "TB AKL AIRPORT",
    "TB NORTHCOTE",
    "TASTE OF INDIA",
    "TOASTED ESPRESSO",
    "VIBE CAFE",
    "VICKIES",
    "WING WONG",
]
DEFAULT_RULES.extend(Rule(text, Classification("支出", "生活费", "外食"), priority=80) for text in FOOD_RULES)


class RuleEngine:
    def __init__(self, rules: list[Rule]) -> None:
        self.rules = sorted(rules, key=lambda rule: rule.priority, reverse=True)

    @classmethod
    def load(cls, local_rules_file: Path) -> "RuleEngine":
        rules = list(DEFAULT_RULES)
        rules.extend(load_local_rules(local_rules_file))
        LOGGER.debug("Loaded %d classification rules", len(rules))
        return cls(rules)

    def classify(self, tx: Transaction) -> Classification | None:
        td = classify_term_deposit(tx)
        if td:
            return td
        if tx.direction == "收入" and "APPLE ONE" in tx.particulars.upper():
            return Classification("收入", "应用软件")
        if tx.direction == "支出" and tx.bank_type in {"BP", "IB", "DC", "AP"} and "GIFT" in tx.particulars.upper():
            return Classification("支出", "人情")
        for rule in self.rules:
            if rule.matches(tx) and is_valid_classification(rule.classification):
                return rule.classification
        return None


def load_local_rules(path: Path) -> list[Rule]:
    if not path.exists():
        return []
    rules: list[Rule] = []
    with path.open(newline="", encoding="utf-8-sig") as handle:
        reader = csv.DictReader(handle)
        for row in reader:
            match_text = str(row.get("match_text") or "").strip()
            tx_type = str(row.get("类型") or "").strip()
            primary = str(row.get("一级分类") or "").strip()
            if not match_text or not tx_type or not primary:
                continue
            amount_text = str(row.get("金额") or "").strip()
            classification = Classification(
                tx_type=tx_type,
                primary=primary,
                secondary=str(row.get("二级分类") or "").strip(),
                account2=str(row.get("账户2") or "").strip(),
            )
            if not is_valid_classification(classification):
                LOGGER.warning("Ignoring invalid local rule: %s -> %s/%s", match_text, tx_type, primary)
                continue
            rules.append(
                Rule(
                    match_text=match_text,
                    classification=classification,
                    priority=int(row.get("优先级") or 200),
                    direction=str(row.get("BNZ方向") or "").strip(),
                    bank_type=str(row.get("BNZ类型") or "").strip().upper(),
                    amount=Decimal(amount_text) if amount_text else None,
                )
            )
    return rules


def classify_term_deposit(tx: Transaction) -> Classification | None:
    if tx.bank_type != "TD" or not tx.other_account:
        return None
    parts = [part for part in tx.other_account.split("-") if part]
    suffix = "-".join(parts[-2:]) if len(parts) >= 2 else tx.other_account
    return Classification("转账", "其他", account2=f"Term Deposit {suffix}")


def is_valid_classification(classification: Classification) -> bool:
    if classification.tx_type == "支出":
        return classification.primary in EXPENSE_CATEGORIES
    if classification.tx_type == "收入":
        return classification.primary in INCOME_CATEGORIES
    if classification.tx_type == "转账":
        return classification.primary == "其他"
    return False


def normalized_unknown_classification(row: dict[str, object]) -> Classification | None:
    row_type = str(row.get("类型") or "").strip()
    primary = str(row.get("一级分类") or "").strip()
    secondary = str(row.get("二级分类") or "").strip()
    bank_direction = str(row.get("BNZ方向") or "").strip()

    if not row_type and primary in SHORT_CATEGORY_MAP:
        primary, secondary = SHORT_CATEGORY_MAP[primary]
        row_type = bank_direction
    elif not row_type and primary:
        row_type = bank_direction
    elif row_type in SHORT_CATEGORY_MAP and not primary:
        primary, secondary = SHORT_CATEGORY_MAP[row_type]
        row_type = bank_direction

    if not row_type and not primary:
        return None

    classification = Classification(
        tx_type=row_type,
        primary=primary,
        secondary=secondary,
        account2=str(row.get("账户2") or "").strip(),
    )
    if not is_valid_classification(classification):
        LOGGER.warning(
            "Skipping invalid manual category: %s %s -> %s/%s",
            row.get("日期"),
            row.get("particulars"),
            classification.tx_type,
            classification.primary,
        )
        return None
    return classification


def read_manual_workbook(path: Path) -> dict[tuple[str, int, str, str, str], Classification]:
    manual: dict[tuple[str, int, str, str, str], Classification] = {}
    if not path.exists():
        return manual
    workbook = load_workbook(path, data_only=True)
    sheet = workbook.active
    headers = [cell.value for cell in sheet[1]]
    for values in sheet.iter_rows(min_row=2, values_only=True):
        row = dict(zip(headers, values))
        if not row.get("particulars"):
            continue
        key = tx_key(row.get("日期"), row.get("金额"), row.get("particulars"), row.get("BNZ类型"), row.get("BNZ方向"))
        classification = Classification(
            tx_type=str(row.get("类型") or ""),
            primary=str(row.get("一级分类") or ""),
            secondary=str(row.get("二级分类") or ""),
            account2=str(row.get("账户2") or ""),
        )
        if is_valid_classification(classification):
            manual[key] = classification
    LOGGER.debug("Loaded %d persisted manual classifications from %s", len(manual), path)
    return manual


def read_filled_unknown_workbook(path: Path) -> dict[tuple[str, int, str, str, str], Classification]:
    manual: dict[tuple[str, int, str, str, str], Classification] = {}
    if not path.exists():
        return manual
    workbook = load_workbook(path, data_only=True)
    sheet = workbook.active
    headers = [cell.value for cell in sheet[1]]
    for values in sheet.iter_rows(min_row=2, values_only=True):
        row = dict(zip(headers, values))
        classification = normalized_unknown_classification(row)
        if not row.get("particulars") or classification is None:
            continue
        key = tx_key(row.get("日期"), row.get("金额"), row.get("particulars"), row.get("BNZ类型"), row.get("BNZ方向"))
        manual[key] = classification
    LOGGER.debug("Loaded %d filled unknown classifications from %s", len(manual), path)
    return manual


def classification_for(tx: Transaction, manual: dict[tuple[str, int, str, str, str], Classification], rules: RuleEngine) -> Classification | None:
    return manual.get(transaction_key(tx)) or rules.classify(tx)
