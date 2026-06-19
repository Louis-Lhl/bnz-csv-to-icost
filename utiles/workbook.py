from __future__ import annotations

import logging
import re
import shutil
import tempfile
import xml.etree.ElementTree as ET
from pathlib import Path
from zipfile import ZIP_DEFLATED, ZipFile

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from .models import Classification, DEFAULT_CURRENCY, ICOST_HEADERS, UNKNOWN_HEADERS, Transaction, money

LOGGER = logging.getLogger(__name__)


def size_columns(ws) -> None:
    for col_idx, column_cells in enumerate(ws.columns, start=1):
        max_len = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, 10), 42)


def write_workbook(path: Path, headers: list[str], rows: list[list[object]], sheet_name: str) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = sheet_name
    sheet.append(headers)
    for row in rows:
        sheet.append(row)
    size_columns(sheet)
    workbook.save(path)
    LOGGER.debug("Wrote workbook %s with %d rows", path, len(rows))


def write_icost_import(path: Path, rows: list[list[object]]) -> None:
    with tempfile.TemporaryDirectory() as tmpdir:
        generated = Path(tmpdir) / "generated.xlsx"
        write_workbook(generated, ICOST_HEADERS, rows, "icost_template")
        convert_inline_strings_to_shared_strings(generated)
        path.parent.mkdir(parents=True, exist_ok=True)
        shutil.move(generated, path)
    LOGGER.info("Wrote %s (%d rows)", path, len(rows))


def write_unknown(path: Path, rows: list[list[object]]) -> None:
    write_workbook(path, UNKNOWN_HEADERS, rows, "unknown")
    LOGGER.info("Wrote %s (%d unknown rows)", path, len(rows))


def convert_inline_strings_to_shared_strings(path: Path) -> None:
    ns = {
        "main": "http://schemas.openxmlformats.org/spreadsheetml/2006/main",
        "rel": "http://schemas.openxmlformats.org/package/2006/relationships",
    }
    ET.register_namespace("", ns["main"])

    with tempfile.TemporaryDirectory() as tmpdir:
        tmp_path = Path(tmpdir)
        with ZipFile(path) as source:
            source.extractall(tmp_path)

        shared_strings: list[str] = []
        shared_index: dict[str, int] = {}
        sheet_path = tmp_path / "xl" / "worksheets" / "sheet1.xml"
        tree = ET.parse(sheet_path)
        root = tree.getroot()

        for cell in root.findall(".//main:c", ns):
            if cell.get("t") != "inlineStr":
                continue
            text_parts = [node.text or "" for node in cell.findall(".//main:t", ns)]
            value = "".join(text_parts)
            if value not in shared_index:
                shared_index[value] = len(shared_strings)
                shared_strings.append(value)
            for child in list(cell):
                if child.tag == f"{{{ns['main']}}}is":
                    cell.remove(child)
            cell.set("t", "s")
            value_node = ET.SubElement(cell, f"{{{ns['main']}}}v")
            value_node.text = str(shared_index[value])

        tree.write(sheet_path, encoding="utf-8", xml_declaration=True)

        shared_root = ET.Element(
            f"{{{ns['main']}}}sst",
            {"count": str(len(shared_strings)), "uniqueCount": str(len(shared_strings))},
        )
        for value in shared_strings:
            si = ET.SubElement(shared_root, f"{{{ns['main']}}}si")
            t = ET.SubElement(si, f"{{{ns['main']}}}t")
            if value != value.strip():
                t.set("{http://www.w3.org/XML/1998/namespace}space", "preserve")
            t.text = value
        ET.ElementTree(shared_root).write(tmp_path / "xl" / "sharedStrings.xml", encoding="utf-8", xml_declaration=True)

        rels_path = tmp_path / "xl" / "_rels" / "workbook.xml.rels"
        rels_tree = ET.parse(rels_path)
        rels_root = rels_tree.getroot()
        if not any((rel.get("Type") or "").endswith("/sharedStrings") for rel in rels_root):
            ids = {
                int(match.group(1))
                for rel in rels_root
                if (match := re.fullmatch(r"rId(\d+)", rel.get("Id") or ""))
            }
            next_id = max(ids or {0}) + 1
            ET.SubElement(
                rels_root,
                f"{{{ns['rel']}}}Relationship",
                {
                    "Id": f"rId{next_id}",
                    "Type": "http://schemas.openxmlformats.org/officeDocument/2006/relationships/sharedStrings",
                    "Target": "sharedStrings.xml",
                },
            )
        rels_tree.write(rels_path, encoding="utf-8", xml_declaration=True)

        content_path = tmp_path / "[Content_Types].xml"
        content_tree = ET.parse(content_path)
        content_root = content_tree.getroot()
        content_ns = "http://schemas.openxmlformats.org/package/2006/content-types"
        if not any(node.get("PartName") == "/xl/sharedStrings.xml" for node in content_root):
            ET.SubElement(
                content_root,
                f"{{{content_ns}}}Override",
                {
                    "PartName": "/xl/sharedStrings.xml",
                    "ContentType": "application/vnd.openxmlformats-officedocument.spreadsheetml.sharedStrings+xml",
                },
            )
        content_tree.write(content_path, encoding="utf-8", xml_declaration=True)

        backup = path.with_suffix(".xlsx.tmp")
        with ZipFile(backup, "w", ZIP_DEFLATED) as target:
            for file_path in sorted(tmp_path.rglob("*")):
                if file_path.is_file():
                    target.write(file_path, file_path.relative_to(tmp_path).as_posix())
        shutil.move(backup, path)


def icost_row(tx: Transaction, classification: Classification) -> list[object]:
    return [
        tx.date_text,
        classification.tx_type,
        money(tx.amount),
        classification.primary,
        classification.secondary or None,
        tx.account,
        classification.account2 or None,
        tx.note,
        DEFAULT_CURRENCY,
        None,
    ]


def transfer_row(tx: Transaction, account2: str | None = None) -> list[object]:
    destination = account2 or tx.transfer_account
    return [
        tx.date_text,
        "转账",
        money(tx.amount),
        "其他",
        None,
        tx.account,
        destination,
        f"{tx.note} | manual_transfer" if account2 else tx.note,
        DEFAULT_CURRENCY,
        None,
    ]


def unknown_row(tx: Transaction) -> list[object]:
    return [
        tx.date_text,
        tx.direction,
        money(tx.amount),
        tx.particulars,
        tx.bank_type,
        None,
        "",
        "",
        "",
        tx.account,
        "",
        tx.note,
        DEFAULT_CURRENCY,
        "",
    ]
